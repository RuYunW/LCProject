# 用于将magic_train.txt转换为【类名|one-hot】
import re
from openpyxl import Workbook

with open('train_macig_in_test.txt', 'rb') as f:
    line = f.read().decode('utf-8')

line = line.split('\n')
line = filter(None, line)
item_name = []
item_NAME_END = []
item_ATK_END = []
item_DEF_END = []
item_COST_END = []
item_DUR_END = []
item_TYPE_END = []
item_PLAYER_CLS_END = []
item_RACE_END = []
item_RARITY_END = []

for i in line:
    item_name.append(re.findall('(.*?) NAME_END', i))
    item_NAME_END.append(re.findall('NAME_END (.*?) ATK_END', i))
    item_ATK_END.append(re.findall('ATK_END (.*?) DEF_END', i))
    item_DEF_END.append(re.findall('DEF_END (.*?) COST_END', i))
    item_COST_END.append(re.findall('COST_END (.*?) DUR_END', i))
    item_DUR_END.append(re.findall('DUR_END (.*?) TYPE_END', i))
    item_TYPE_END.append(re.findall('TYPE_END (.*?) PLAYER_CLS_END', i))
    item_PLAYER_CLS_END.append(re.findall('PLAYER_CLS_END (.*?) RACE_END', i))
    item_RACE_END.append(re.findall('RACE_END (.*?) RARITY_END', i))
    item_RARITY_END.append(re.findall('RARITY_END (.*?) \.', i))

for i in item_RARITY_END:
    print(i)
wb = Workbook()
ws = wb.active
ws.title = 'sheet1'

# 写抬头
for i in range(len(item_name)):
    ws.cell(row=i + 2, column=1).value = str(item_name[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=2).value = str(item_NAME_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=3).value = str(item_ATK_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=4).value = str(item_DEF_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=5).value = str(item_COST_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=6).value = str(item_DUR_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=7).value = str(item_TYPE_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=8).value = str(item_PLAYER_CLS_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=9).value = str(item_RACE_END[i]).replace('[\'','').replace('\']','')
    ws.cell(row=i + 2, column=10).value = str(item_RARITY_END[i]).replace('[\'','').replace('\']','').replace('[\"','').replace('\"]','')

for i in range(9):
    ws.cell(row=1, column=2).value = 'NAME_END'
    ws.cell(row=1, column=3).value = 'ATK_END'
    ws.cell(row=1, column=4).value = 'DEF_END'
    ws.cell(row=1, column=5).value = 'COST_END'
    ws.cell(row=1, column=6).value = 'DUR_END'
    ws.cell(row=1, column=7).value = 'TYPE_END'
    ws.cell(row=1, column=8).value = 'PLAYER_CLS_END'
    ws.cell(row=1, column=9).value = 'RACE_END'
    ws.cell(row=1, column=10).value = 'RARITY_END'

# 保存
wb.save(filename='magic_train_in_test.xlsx')

