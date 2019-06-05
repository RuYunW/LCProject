import re
import os
from openpyxl import Workbook

filename = 'train_magic_class_method2.txt'
with open(filename, 'rb') as f:
    content = f.read().decode('utf-8')
line = content.split('\r')


with open("train_magic_new_name5.txt", 'rb') as ff:
    new_name = ff.read().decode('utf-8')
    new_name.split("\n")
# replace(re.findall(***),"new_name[i]")!!!!!!!!!!!!!!!!!!!!!!!!!!!
class_list = []
method_list = []
class_name = ""
method_name = ""
class_method_all = []
input = ""
output = ""

# 创建一个worksheet
wb = Workbook()
ws = wb.active
ws.title = "sheet1"
print("文件已创建")

    # ws.cell(row=i + 2, column=1).value = str(node_all[i])  # 写抬头列
    # ws.cell(row=1, column=i + 2).value = str(node_all[i])  # 写抬头行

id = 0
row = 1
method_name_old = ""
for i in line:
    class_name = str(re.findall("class (.*?) {",i)).split(" ")[0]
    class_list.append(class_name.replace("['","")+":")
    method_name = str(re.findall("{(.*?)\(.*?\) {",i)).split(" ")[-1]
    method_list.append(method_name.replace("']","").replace("['",""))
    input = str(re.findall("\((.*?)\) {",i))
    output = str(re.findall("{(.*?)\(.*?\) {",i.replace("public ","").replace("private ",""))).split(" ")[0]

    class_method_all.append(class_name.replace("['","")+"."+method_name.replace("']","").replace("['","")+"-input:("+input.replace("[","").replace("]","").replace("'","")+")-output:("+output.replace("[","").replace("]","").replace("'","")+")")
    ws.cell(row = row,column=1).value = class_name.replace("['","")+"."+method_name.replace("']","").replace("['","")+"()"
    ws.cell(row = row,column=2).value = input.replace("[","").replace("]","").replace("'","").replace("final ","")
    ws.cell(row = row,column=3).value = output.replace("[","").replace("]","").replace("'","")
    row+=1

# 写之前，先检验文件是否存在，存在就删掉
if os.path.exists("train_magic_class_method_io3.txt"):
    os.remove("train_magic_class_method_io3.txt")
if os.path.exists("train_magic_class_name3.txt"):
    os.remove("train_magic_class_name3.txt")

# 以写的方式打开文件，如果文件不存在，就会自动创建
file_write_obj = open("train_magic_class_method_io3.txt", 'w')
for var in class_method_all:
    file_write_obj.writelines(var)
    file_write_obj.write('\n')
file_write_obj.close()

file_write_obj = open("train_magic_class_name3.txt", 'w')
for var in class_list:
    file_write_obj.writelines(var)
    file_write_obj.write('\n')
file_write_obj.close()


print("文件写入完毕")
# 保存
wb.save(filename="magic_io_3.xlsx")
print("文件已保存")





