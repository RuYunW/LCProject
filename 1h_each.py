import numpy as np
from numpy import array,argmax
from sklearn.preprocessing import LabelEncoder,OneHotEncoder
from openpyxl import Workbook

# 打开文件
with open('train_macig_in_test.txt', 'rb') as f:
    line = f.read().decode('utf-8')
line = line.split('\n')
line = list(filter(None, line))
print("文件打开成功")
# line2 = line[:]
# 词典
words = []
for i in line:
    for j in i.split(' '):
        words.append(j)
print("词典建立成功")
# 词 one-hot 编码
print("one-hot编码中……")
values = array(words)
label_encoder = LabelEncoder()
integer_encoded = label_encoder.fit_transform(values)
onehot_encoder = OneHotEncoder(sparse=False)
integer_encoded = integer_encoded.reshape(len(integer_encoded), 1)
onehot_encoded = onehot_encoder.fit_transform(integer_encoded)
inverted = label_encoder.inverse_transform([argmax(onehot_encoded[0, :])])
print("编码完毕")


# 写onehot
# for i in range(len(onehot_encoded)):
#     ws.cell(row=i+1,column=1).value = str(words[i])
#     ws.cell(row=i+1,column=2).value = str(onehot_encoded[i])
#
# wb.save('word_one_hot.xlsx')

# print(type(onehot_encoded[0]))
# print(type(np.array(np.zeros(len(words)))))
print("数据计算中……")
ont_hot_seq = []
for i in line:
    temp = onehot_encoded[0]-onehot_encoded[0]
    for j in i.split(' '):
        temp = temp + onehot_encoded[words.index(str(j))]
    ont_hot_seq.append(temp)



print("数据计算完毕")

wb = Workbook()
ws = wb.active
ws.title = 'sheet1'
print("文件创建成功")

print("文件写入中……")
for i in range(len(line)):
    for j in range(len(onehot_encoded[0])):
        ws.cell(row=i+1,column=j+1).value = str(ont_hot_seq[i][j])


wb.save('final_one_hot_encode.xlsx')
print("文件写入结束")
