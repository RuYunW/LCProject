'''
1. 清洗method文件数据
2. 分离input output class
'''

import os
import re
from openpyxl import load_workbook, Workbook
'''
1. 清洗
'''
def readfile(filename):
    inlist = []
    outlist = []
    method_list = []
    with open(filename,'r') as f:
        lines = f.readlines()
        for line in lines:
            method = str(line).replace('\r','').replace('public ','').replace('static ','').replace('@Override ','').replace('private ','').replace('@Deprecated ','').replace('final ','').replace('\n','').replace('@deprecated ','').replace('protected ','').replace('@ShortType ','').replace('@VisibleForTesting ','')
            # method = re.sub(r'@SuppressWarnings\({".*"}\) ','',method)
            # # method = re.sub(r'@SuppressLint("(.*)") ','',method)
            # method = re.sub(r'@SuppressLint\(".*"\) ','',method)
            # method = re.sub(r'@PostConstruct ','',method)
            method = re.sub('@(.*) ','',method)
            # print(method)
            input = re.findall(r' .*?\((.*?)\){',method)
            output = re.findall(r'(.*?) .*?\(.*?\){',method)
            input_class = []
            for i in str(input).split(','):
                input_class.append(i.split(' ')[0].replace('[','').replace(']','').replace('\'',''))
            inlist.append(input_class)
            outlist.append(output)
            method_name = re.findall(r'.*? (.*?)\(.*?\)',method)
            method_list.append(method_name)
            # print(input)
            # print(output)
    return inlist,outlist,method_list
        # return lines






filename = './data/all_method.txt'
inlist,outlist,method_list = readfile(filename)

# filename2 = "./data/method_io.xlsx"
def writefile(input,output,method):
    # 创建一个worksheet
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "sheet1"

    print("文件已创建")
    for i in range(len(method)):
        ws1.cell(row=i+1,column=1).value = str(method[i])
        ws1.cell(row=i+1,column=2).value = str(input[i])
        ws1.cell(row=i+1,column=3).value = str(output[i])
    print("文件写入完毕")
    # 保存
    wb1.save(filename="./data/method_io.xlsx")

    print("文件已保存")

writefile(inlist,outlist,method_list)