from openpyxl import Workbook, load_workbook

# 打开文件：
excel = load_workbook('magic_io_final_test.xlsx')
excel1 = load_workbook('magic_input_4_test.xlsx')
excel2 = load_workbook('magic_output_4_test.xlsx')

# 获取sheet：
table = excel.worksheets[0]
table1 = excel1.worksheets[0]
table2 = excel2.worksheets[0]

node_all = []



# 获取所有method node
for i in range(table.max_row):
    node_all.append(table.cell(row=i + 1, column=1).value)

n_input = []
method_input = []
# 获取input
for i in range(table1.max_row):
    temp = table1.cell(row=i + 1, column=1).value
    n_input.append(temp)
    method_input.append(str(table1.cell(row=i + 1, column=2).value).split(','))

n_output = []
method_output = []
# 获取output
for i in range(table2.max_row):
    temp = table2.cell(row=i + 1, column=1).value
    n_output.append(temp)
    method_output.append(str(table2.cell(row=i + 1, column=2).value).split(','))

# 创建一个worksheet
wb = Workbook()
ws = wb.active
ws.title = "sheet1"
print("文件已创建")

for i in range(len(node_all)):
    ws.cell(row=i + 2, column=1).value = str(node_all[i])  # 写抬头列
    ws.cell(row=1, column=i + 2).value = str(node_all[i])  # 写抬头行

for index_output in range(len(n_output)):
    temp = n_output[index_output]
    if temp in n_input:
        index_input = n_input.index(temp)
        for i in range(len(method_output[index_output])):
            for j in range(len(method_input[index_input])):
                # print(method_input[index_input][j])
                if method_output[index_output][i] in node_all and method_input[index_input][j] in node_all:
                    # print(666)
                    matrix_index_i = node_all.index(method_output[index_output][i])
                    matrix_index_j = node_all.index(method_input[index_input][j])
                    # c = ws.cell(row=matrix_index_i, column=matrix_index_j)
                    ws.cell(row=matrix_index_i + 2, column=matrix_index_j + 2).value = '1'

                    # print(matrix_index_i,matrix_index_j)
                    # node_matrix[matrix_index_i][matrix_index_j] = 1

                    # worksheet.write(matrix_index_j, matrix_index_j, label=1)
print("文件写入完毕")

# 保存
wb.save(filename="magic_fig_node_0-1.xlsx")
print("文件已保存")
