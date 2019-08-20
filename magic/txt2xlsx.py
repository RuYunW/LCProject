import re
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "sheet1"
print("文件已创建")

with open('magic_final_result.txt','rb') as f:
    line = f.read().decode('utf-8')
line = line.split('\r')

method_name = []
input = []
output = []

for i in range(len(line)):
    method_name.append(str(line[i]).split('-')[0])
    input.append(re.findall("input:\((.*?)\)",str(line[i])))
    output.append(re.findall("output:\((.*?)\)", str(line[i])))
    ws.cell(row=i+1,column=1).value = str(method_name[i])
    ws.cell(row=i+1,column=2).value = str(input[i])
    ws.cell(row=i+1,column=3).value = str(output[i])

wb.save(filename='magic_io_final.xlsx')
