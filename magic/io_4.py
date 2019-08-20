from openpyxl import load_workbook, Workbook

# 打开文件：
# excel = load_workbook('magic_io_3_test.xlsx')
excel = load_workbook('magic_io_final_test.xlsx')
# 获取sheet：
table = excel.worksheets[0]
# 获取行数和列数：
rows = table.max_row  # 获取行数
cols = table.max_column  # 获取列数

input = []
output = []

for i in range(1, rows + 1):
    for j in str(table.cell(row=i, column=2).value).split(","):
        for k in j.split(" "):
            if len(k) >= 2:
                input.append(str(k))
    for p in str(table.cell(row=i, column=3).value).split(","):
        for q in p.split(" "):
            if len(q) >= 2:
                output.append(str(q))
    # print(i)

input = list(set(input))
output = list(set(output))

# 创建一个worksheet
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "sheet1"
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "sheet1"
print("文件已创建")
input_line = ""
output_line = ""

print(len(input))
print(len(output))
for j in range(len(input)):
    for i in range(1, rows + 1):
        if str(input[j]) in str(table.cell(row=i, column=2).value):
            input_line += str(table.cell(row=i, column=1).value)+","
    ws1.cell(row=j + 1, column=2).value = input_line[0:-1]
    ws1.cell(row = j+1,column=1).value = input[j].replace('[','').replace(']','').replace('\'','')
    input_line = ""
    print(j)

for k in range(len(output)):
    for i in range(1, rows + 1):
        if str(output[k]) in str(table.cell(row=i, column=3).value):
            output_line += str(table.cell(row=i, column=1).value) + ","
    ws2.cell(row=k + 1, column=2).value = output_line[0:-1]
    ws2.cell(row = k+1,column=1).value = output[k].replace('[','').replace(']','').replace('\'','')
    output_line = ""
    print(k)



print("文件写入完毕")
# 保存
wb1.save(filename="magic_input_4_test.xlsx")
wb2.save(filename="magic_output_4_test.xlsx")
print("文件已保存")

