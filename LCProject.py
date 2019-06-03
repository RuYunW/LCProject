import csv
import sys
import re
import numpy as np
import xlwt

from openpyxl import Workbook

# 大文件读取声明
maxInt = sys.maxsize
decrement = True
while decrement:
    decrement = False
    try:
        csv.field_size_limit(maxInt)
    except OverflowError:
        maxInt = int(maxInt / 10)
        decrement = True

# 文件路径
path_input = 'result_input.csv'
path_output = 'result_output.csv'

# 存储
line_input = []  # 行
line_output = []
node = []  # 节点

# 读input文件
with open(path_input) as f:
    reader_input = csv.reader(f)
    # 遍历
    for index, i in enumerate(reader_input):  # 原始行
        if index != 0:  # 去除标题列
            line_input.append(i)  # input行
print("input文件读取完成")

# 得到input所有方法
node_input = str(np.array(line_input[:])[:, 1])
node_input = node_input.replace("[", "").replace("]", "").replace("\'", "").replace("\"", "").replace(" ", "").replace(
    "\n", ",")  # 清洗
node_input = node_input.split(",")

# 读output文件
with open(path_output) as ff:
    reader_output = csv.reader(ff)
    # 遍历
    for index, i in enumerate(reader_output):
        if index != 0:  # 去除标题列
            line_output.append(i)  # output行
print("output文件读取完成")

# 得到output所有方法
node_output = str(np.array(line_output[:])[:, 1])
node_output = node_output.replace("[", "").replace("]", "").replace("\'", "").replace("\"", "").replace("\n",
                                                                                                        ",").replace(
    " ", "")  # 清洗
node_output = node_output.split(",")

node_all = node_input + node_output  # 连接input节点与output节点
node_all = list(set(node_all))  # 转集合去重
print("节点获取完毕")
# node_matrix = np.zeros((len(node_all), len(node_all)))  # 初始化图边矩阵

# output
n_output = []
for i in line_output:
    n_output.append(i[0])

# input
n_input = []
for i in line_input:
    n_input.append(i[0])

method_output = []
method_input = []
for i in range(len(n_output)):
    for j in range(len(n_input)):
        method_output.append(
            line_output[i][1].replace("\'", "").replace("[", "").replace("]", "").replace(" ", "").replace("\n",
                                                                                                           ",").split(
                ","))
        method_input.append(
            line_input[j][1].replace("\'", "").replace("[", "").replace("]", "").replace(" ", "").replace("\n",
                                                                                                          ",").split(
                ","))

# # 创建一个worksheet
wb = Workbook()
ws = wb.active
ws.title = "sheet1"
print("文件已创建")

for i in range(len(node_all)):
    ws.cell(row=i + 2, column=1).value = str(node_all[i])  # 写抬头列
    ws.cell(row=1, column=i + 2).value = str(node_all[i])  # 写抬头行

for index_output in range(len(n_output)):
    temp = n_output[index_output]
    if temp in n_input:
        index_input = n_input.index(temp)
        for i in range(len(method_output[index_output])):
            for j in range(len(method_input[index_input])):
                # print(method_input[index_input][j])
                if method_output[index_output][i] in node_all and method_input[index_input][j] in node_all:
                    # print(666)
                    matrix_index_i = node_all.index(method_output[index_output][i])
                    matrix_index_j = node_all.index(method_input[index_input][j])
                    # c = ws.cell(row=matrix_index_i, column=matrix_index_j)
                    ws.cell(row=matrix_index_i + 2, column=matrix_index_j + 2).value = '1'

                    # print(matrix_index_i,matrix_index_j)
                    # node_matrix[matrix_index_i][matrix_index_j] = 1

                    # worksheet.write(matrix_index_j, matrix_index_j, label=1)
print("文件写入完毕")

# # 保存
wb.save(filename="fig_node_0-1.xlsx")
print("文件已保存")
